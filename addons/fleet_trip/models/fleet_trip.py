# -*- coding: utf-8 -*-
from dateutil.relativedelta import relativedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError
import base64
from openpyxl.drawing.image import Image
import requests
import openpyxl
from io import BytesIO
from openpyxl.utils import units
from odoo.modules.module import get_module_resource
from openpyxl.styles import Alignment
from datetime import datetime
import re

def capitalize_first_letter(s):
    # Kiểm tra nếu chuỗi là False, None hoặc rỗng
    if not s:
        return s
    # Viết hoa chữ cái đầu tiên và giữ nguyên các ký tự còn lại
    return s[0].upper() + s[1:]

def remove_duplicate_adjacent_words(sentence):
    # This regex will match any word (\\w+) that is followed by the same word (\\1)
    return re.sub(r'\b(\w+)\s+\1\b', r'\1', sentence)

class FleetTrip(models.Model):
    _name = 'fleet.trip'
    _rec_name = 'equipment_id'
    _order = 'schedule_date asc'
    _description = 'Hành trình vận tải'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    # def _get_location_selection(self):
    #     selection = []
    #     list_location = self.env['fleet.location'].search([])
    #     for location in list_location:
    #         selection += [(location.code, location.name)]
    #     return selection
    fleet_code = fields.Integer(string='Số Dự trù')
    fleet_command_code = fields.Integer(string='Mã Lệnh')
    fleet_preventive = fields.Integer(string='Số lượng dự phòng')
    company_id = fields.Many2one('res.company', 'Company', default=lambda self: self.env.company)
    department_id = fields.Many2one('hr.department', string="phòng ban")
    acronym_department = fields.Char(related='department_id.acronym',string='Tên viết tắt phòng ban')
    department_plan_id = fields.Many2one('hr.department', string="Đơn vị dự trù phương tiện",  default=lambda self: self._default_department())
    acronym_department_plan = fields.Char(related='department_plan_id.acronym',string='Tên viết tắt phòng ban dự trù')
    department_belong_id = fields.Many2one(related='equipment_id.department_belong_id', string="Thuộc đơn vị")
    currency_id = fields.Many2one('res.currency', related='company_id.currency_id')
    equipment_id = fields.Many2one('maintenance.equipment', string='Xe')
    model_id = fields.Many2one('fleet.vehicle.model',related='equipment_id.model_id', string="Nhãn xe")
    category_plan = fields.Many2one('maintenance.equipment.category', string='Tên phương tiện')
    category_plan_name = fields.Char(related='category_plan.name', string='Tên phương tiện')
    vehicle_id = fields.Many2one(related='equipment_id.vehicle_id', string='Phương tiện')
    license_plate = fields.Char(related='vehicle_id.license_plate', string='Biển số')
    fuel_id = fields.Many2one(related='equipment_id.fuel_id', string='Nhiên liệu')
    location_name = fields.Char()
    location_dest_name = fields.Char()
    # location_id = fields.Selection(selection=_get_location_selection)
    # location_dest_id = fields.Selection(selection=_get_location_selection)
    location_id = fields.Many2one('fleet.location', string="Tên điểm đi")
    location_dest_id = fields.Many2one('fleet.location', string="Tên điểm đến")
    eating_fee = fields.Monetary('Tiền ăn')
    law_money = fields.Monetary('Tiền luật')
    road_tiket_fee = fields.Monetary('Vé cầu đường')
    incurred_fee = fields.Monetary('Phát sinh')
    incurred_note = fields.Char('Ghi chú phát sinh')
    incurred_fee_2 = fields.Monetary('Phát sinh 2')
    incurred_note_2 = fields.Char('Ghi chú phát sinh 2')
    note = fields.Text('Ghi chú hành trình')
    fee_total = fields.Monetary('Tổng cộng', compute='_compute_fee_total')
    quota = fields.Integer(related='equipment_id.quota',string='Định mức')
    number_picks = fields.Integer('Người đón xe')
    number_trips = fields.Integer('Số Chuyến')
    # attemdances = fields.One2many('hr.employee', 'id',
    #                                 #  domain=[('res_model', '=', 'fleet.trip')],
    #                                  string='Người tham gia')
    # attendances = fields.One2many('hr.employee', 'id',
    #                                 #  domain=[('res_model', '=', 'fleet.trip')],
    #                                  string='Người tham gia')

    employee_ids = fields.Many2many('hr.employee',
                                     string='Người tham gia')
    number_people = fields.Integer('Số người')
    product_weigh = fields.Float(string='Số tấn HH')
    number_seat = fields.Integer(related='equipment_id.number_seat',string='Số ghế')
    odometer_start = fields.Integer('Số CTM xuất phát')
    odometer_dest = fields.Integer('Số CTM điểm đích')
    odometer_end = fields.Integer('Số KM hành trình', compute='_compute_odometer_end', store=True)
    distance_plan = fields.Integer('Dự kiến tổng số km đi, về')
    route = fields.Many2one('fleet.route', string='Cung đường')
    employee_plan_id = fields.Many2one('hr.employee', string='Người dự trù', default=lambda self: self._default_employee())
    employee_lead_id = fields.Many2one('hr.employee', string='Chỉ huy xe',
                                        domain="[('id', 'in', employee_ids)]")
    employee_command_id = fields.Many2one('hr.employee', string='Người viết lệnh')
    employee_approved_id = fields.Many2one('hr.employee', string='Người duyệt lệnh', domain="[('job_id.is_manage_company', '=', True)]", default=lambda self: self._default_employee_approved())
    
    level = fields.Char( string='Cấp bậc')
    position = fields.Char( string='Chức vụ')
    job_id = fields.Many2one("hr.job", related='employee_lead_id.job_id', string='Chức vụ cán bộ')
    rank_id = fields.Many2one("hr.rank",related='employee_lead_id.rank_id', string='Cấp bậc cán bộ')
    employee_id = fields.Many2one('hr.employee', string='Nhân viên')
    state = fields.Selection([
        ('0_plan', 'Khởi tạo'),
        ('1_draft', 'Dự trù'),
        ('2_command', 'Lệnh điều phương tiện'),
        ('2_confirm', 'Đã Xuất Phát'),
        ('3_done', 'Hoàn Thành')
    ], string='Trạng thái', default='0_plan')
    schedule_date = fields.Datetime(string='Ngày Giờ Dự kiến')
    start_date = fields.Datetime(string='Giờ Bắt đầu', readonly=False)
    end_date = fields.Datetime(string='Giờ Kết thúc', readonly=False)
    time_day_compute = fields.Float('Tổng cộng', compute='_compute_timeday')
    delivery_id = fields.Many2one('stock.delivery', string='Phiếu xuất kho')
    code = fields.Char(related='delivery_id.code', store=True)
    project_id = fields.Many2one(related='delivery_id.project_id')

    district_id = fields.Many2one('res.country.district', string='Huyện', domain="[('state_id', '=', state_id)]")
    ward_id = fields.Many2one('res.country.ward', string='Xã', domain="[('district_id', '=', district_id)]")
    state_id = fields.Many2one("res.country.state", string='Tỉnh', ondelete='restrict',
                               domain="[('country_id', '=', country_id)]")

    location_start_district_state = fields.Char(string='Địa chỉ điểm đi',
                                                compute='_compute_location_start_district_state')
    location_dest_district_state = fields.Char(string='Địa chỉ điểm đến',
                                               compute='_compute_location_dest_district_state')

    location_compute_name = fields.Char(string='Nơi xuất phát',
                                        compute='_compute_location_compute_name')
    location_dest_compute_name = fields.Char(string='Nơi đến',
                                             compute='_compute_location_dest_compute_name')

    district_dest_id = fields.Many2one('res.country.district', string='Huyện',
                                       domain="[('state_id', '=', state_dest_id)]")
    ward_dest_id = fields.Many2one('res.country.ward', string='Xã', domain="[('district_id', '=', district_dest_id)]")
    state_dest_id = fields.Many2one("res.country.state", string='Tỉnh', ondelete='restrict',
                                    domain="[('country_id', '=', country_id)]")

    country_id = fields.Many2one('res.country', default=241, string='Quốc gia', ondelete='restrict')
    company_name = fields.Char(string='Công ty')
    fleet_product_id = fields.Many2one('fleet.product', string='Mặt hàng', ondelete='restrict')
    address_start = fields.Char(string="Địa chỉ xuất phát")
    address_end = fields.Char(string="Địa chỉ đích")
    start_hour = fields.Datetime(string="Giờ xuất phát")
    end_hour = fields.Datetime(string="Giờ đến đích")
    is_approved = fields.Boolean(string="Đã xác nhận")
    attachment_ids = fields.One2many('ir.attachment', 'res_id',
                                     domain=[('res_model', '=', 'fleet.trip')],
                                     string='Attachments')
    description = fields.Text(string='Nhiệm vụ')

    def _default_employee(self):
        return self.env.user.employee_id

    def _default_department(self):
        employee = self.env.user.employee_id
        return employee.department_id if employee else False
    
    @api.model 
    def _default_employee_approved(self):
        # Get the first employee who is a company manager 
        employee = self.env['hr.employee'].search([('job_id.is_manage_company', '=', True)], limit=1) 
        return employee.id if employee else False

    @api.model
    def create(self, vals):
        max_item = self.search([('department_plan_id','=', vals['department_plan_id'])], order='fleet_command_code desc', limit=1)
        max_value = max_item.fleet_command_code
        vals['fleet_command_code'] = max_value + 1 if max_value else 1
        max_item = self.search([('department_plan_id','=', vals['department_plan_id'])], order='fleet_code desc', limit=1)
        fleet_code_max = max_item.fleet_code
        vals['fleet_code'] = fleet_code_max + 1 if fleet_code_max else 1
        return super(FleetTrip, self).create(vals)
    # @api.model
    # def create(self, vals):

    # fleet_preventive = vals.get('fleet_preventive')
    # if fleet_preventive and (fleet_preventive > 0):
    #     for i in range(0, fleet_preventive):
    #         newval = {}
    #         newval['fleet_preventive'] = 0
    #         self.create(newval)

    # result = super(FleetTrip, self).create(vals)
    # return result

    # @api.model
    def write(self, vals):
        old_state = self.state
        new_state = vals.get('state')
        if (new_state == '1_draft') and (old_state != new_state):
            fleet_preventive = vals.get('fleet_preventive')
            if fleet_preventive and (fleet_preventive > 0):
                for i in range(0, fleet_preventive):
                    newval = {}
                    newval['fleet_preventive'] = 0
                    self.create(newval)

        res = super(FleetTrip, self).write(vals)
        return res

    @api.onchange("location_id")
    def onchange_location_id(self):
        if self.location_id:
            self.location_name = self.location_id.name
            self.district_id = self.location_id.district_id.id
            self.ward_id = self.location_id.ward_id.id
            self.state_id = self.location_id.state_id.id
            self.address_start = self.location_id.note

    @api.onchange("location_dest_id")
    def onchange_location_dest_id(self):
        if self.location_dest_id:
            self.location_dest_name = self.location_dest_id.name
            self.district_dest_id = self.location_dest_id.district_id.id
            self.ward_dest_id = self.location_dest_id.ward_id.id
            self.state_dest_id = self.location_dest_id.state_id.id
            self.address_end = self.location_dest_id.note

    @api.depends("district_id", "ward_id", "state_id")
    def _compute_location_compute_name(self):
        for record in self:
            location_name = []
            if record.ward_id:
                location_name.append(record.ward_id.name or '')
            if record.district_id:
                location_name.append(record.district_id.name or '')
            if record.state_id:
                location_name.append(record.state_id.name or '')
            record.location_compute_name = ', '.join(location_name)

    @api.depends("district_dest_id", "ward_dest_id", "state_dest_id")
    def _compute_location_dest_compute_name(self):
        for record in self:
            location_name = []
            if record.ward_dest_id:
                location_name.append(record.ward_dest_id.name or '')
            if record.district_dest_id:
                location_name.append(record.district_dest_id.name or '')
            if record.state_dest_id:
                location_name.append(record.state_dest_id.name or '')
            record.location_dest_compute_name = ', '.join(location_name)

    @api.onchange("equipment_id")
    def _onchange_equipment_id(self):
        if self.equipment_id and self.equipment_id.owner_user_id and self.equipment_id.owner_user_id.employee_id:
            self.employee_id = self.equipment_id.owner_user_id.employee_id.id

    @api.onchange("employee_id")
    def _onchange_employee_id(self):
        if self.employee_id and self.employee_id.user_id:
            equipment_id = self.env['maintenance.equipment'].search([
                ('owner_user_id', '=', self.employee_id.user_id.id)], limit=1)
            if equipment_id:
                self.equipment_id = equipment_id.id

    @api.depends("eating_fee", "law_money", "road_tiket_fee", "incurred_fee", "incurred_fee_2")
    def _compute_fee_total(self):
        for rec in self:
            rec.fee_total = rec.eating_fee + rec.law_money + rec.road_tiket_fee + rec.incurred_fee + rec.incurred_fee_2

    def custom_round_half_day(self, start_date, end_date):
        rounded_days = 0
        if (start_date != False) and (end_date!=False):
            delta = end_date - start_date
            total_days = (delta.total_seconds() / (24.0 * 3600)) % 10
            floor_days = delta.total_seconds() // (24 * 3600)

            rounded_days = floor_days if total_days == 0 else floor_days + 0.5
            rounded_days = rounded_days + 0.5 if total_days > 0.5 else rounded_days

        return rounded_days

    @api.depends("start_date", "end_date")
    def _compute_timeday(self):
        for rec in self:
            rec.time_day_compute = self.custom_round_half_day(self.start_date, self.end_date)

    def do_plan_trip(self):
        # self.start_date = fields.Datetime.now()
        self.state = '1_draft'

    def do_start_trip(self):
        self.start_date = fields.Datetime.now()
        self.state = '2_confirm'

    def do_create_command_trip(self):
        self.start_date = fields.Datetime.now()
        self.state = '2_command'

    def do_end_trip(self):
        self.end_date = fields.Datetime.now()
        self.state = '3_done'

    def do_odometer_start(self, odometer_start, attachments=[]):
        self.odometer_start = odometer_start
        if not attachments:
            return True
        for attachment in attachments:
            self.env['ir.attachment'].create({
                'name': self.equipment_id.name,
                'type': 'url',
                'url': attachment,
                'res_model': 'fleet.trip',
                'res_id': self.id,
            })

    def do_odometer_end(self, odometer_end, attachments=[]):
        self.odometer_dest = odometer_end
        if not attachments:
            return True
        for attachment in attachments:
            self.env['ir.attachment'].create({
                'name': self.equipment_id.name,
                'type': 'url',
                'url': attachment,
                'res_model': 'fleet.trip',
                'res_id': self.id,
            })

    def do_odometer_dest(self, odometer_dest, attachments=[]):
        self.odometer_dest = odometer_dest
        if not attachments:
            return True
        for attachment in attachments:
            self.env['ir.attachment'].create({
                'name': self.equipment_id.name,
                'type': 'url',
                'url': attachment,
                'res_model': 'fleet.trip',
                'res_id': self.id,
            })

    @api.depends('odometer_start', 'odometer_dest')
    def _compute_odometer_end(self):
        for record in self:
            odometer_total = record.odometer_dest - record.odometer_start
            record.odometer_end = odometer_total if odometer_total >= 0 else 0

    @api.depends("district_id", "state_id")
    def _compute_location_start_district_state(self):
        for record in self:
            location_name = []
            if record.district_id:
                location_name.append(record.district_id.name or '')
            if record.state_id:
                location_name.append(record.state_id.name or '')
            record.location_start_district_state = ', '.join(location_name)

    @api.depends("district_dest_id", "state_dest_id")
    def _compute_location_dest_district_state(self):
        for record in self:
            location_name = []
            if record.district_dest_id:
                location_name.append(record.district_dest_id.name or '')
            if record.state_dest_id:
                location_name.append(record.state_dest_id.name or '')
            record.location_dest_district_state = ', '.join(location_name)

    # @api.onchange('location_id', 'location_dest_id')
    # def onchange_location(self):
    #     company_name = ''
    #     if self.location_id:
    #         company_name = "NM" + self.location_id
    #     elif self.location_dest_id:
    #         company_name = "NM" + self.location_dest_id
    #     self.company_name = company_name

    def do_approve(self):
        view_id = self.env.ref('fleet_trip.fleet_trip_approve_form_view').id
        act_window = {
            'type': 'ir.actions.act_window',
            'name': (_("Confirmation")),
            'res_model': 'fleet.trip.approve.reject',
            'view_mode': 'form',
            'view_type': 'form',
            'view_id': view_id,
            'views': [(view_id, 'form')],
            'target': 'new',
            'context': {'default_fleet_trip_id': self.id}}
        return act_window

    def do_reject(self):
        view_id = self.env.ref('fleet_trip.fleet_trip_reject_form_view').id
        act_window = {
            'type': 'ir.actions.act_window',
            'name': (_("Confirmation")),
            'res_model': 'fleet.trip.approve.reject',
            'view_mode': 'form',
            'view_type': 'form',
            'view_id': view_id,
            'views': [(view_id, 'form')],
            'target': 'new',
            'context': {'default_fleet_trip_id': self.id}}
        return act_window

    def action_download_command_template(self):
        # current time
        now = datetime.now()
        # Load the template
        file_path = get_module_resource('fleet_trip', 'static/src/template', 'COMMAND_TEMPLATE.xlsx')
        workbook = openpyxl.load_workbook(file_path)

        # Access the worksheets
        ws1 = workbook['Lệnh']
        #update ma lệnh
        ws1.cell(row=4, column=3).value = f"{self.fleet_command_code:02}" if self.fleet_command_code else ''
        ws1.cell(row=4, column=8).value = f"Hà Nội, ngày {now.day:02} tháng {now.month:02} Năm {now.year}"
        ws1.cell(row=8, column=10).value = f"{self.fleet_code:02}" if self.fleet_code else ''
        ws1.cell(row=8, column=13).value = f"{self.acronym_department_plan or ''}"
        ws1.merge_cells(start_row=8, start_column=13, end_row=8, end_column=16)

        # loại xe
        ws1.cell(row=9, column=5).value = (f"{self.category_plan_name or ''}") 
        ws1.merge_cells(start_row=9, start_column=5, end_row=9, end_column=9)

        # nhan xe
        
        ws1.cell(row=9, column=12).value = (f"{self.model_id.name or ''}") if self.model_id else ''
        ws1.merge_cells(start_row=9, start_column=12, end_row=9, end_column=16)
        
        # department_belong_id
        ws1.cell(row=10, column=6).value = (f"{self.department_belong_id.name or ''}") if self.department_belong_id else ''
        ws1.merge_cells(start_row=10, start_column=6, end_row=10, end_column=16)
        
        ws1.cell(row=11, column=8).value = (f"{self.license_plate}") if self.license_plate else ''
        ws1.merge_cells(start_row=11, start_column=8, end_row=11, end_column=16)
        # nhiệm vụ
        ws1.cell(row=12, column=8).value = (
            {self.description} if self.description
                else ""
        )        
        # số chuyến
        ws1.cell(row=13, column=14).value = f"{self.number_trips:02d}" if (self.number_trips) else ""
        # cung đường    
        path = ""
        equipment_location_name = self.equipment_id.location_id.name
        # Construct the path based on the provided locations
        if equipment_location_name == self.location_name:
            path = f"{equipment_location_name}-{self.location_dest_name}-{equipment_location_name}"
        elif equipment_location_name == self.location_dest_name:
            path = f"{equipment_location_name}-{self.location_name}-{equipment_location_name}"
        else:
            path = f"{equipment_location_name}-{self.location_name},{self.location_dest_name}-{equipment_location_name}"
        
        # Write the constructed path to the specified cell
        ws1.cell(row=14, column=6).value = path
        # thời gian
        ws1.cell(row=16, column=2).value = f"- Thời gian: {self.time_day_compute} ngày;" if (self.start_date) else ''
        ws1.cell(row=16, column=8).value = f"{self.start_date.hour:02d}" if (self.start_date) else ''
        if (self.start_date):
            if self.start_date.minute > 0:
                ws1.cell(row=16, column=10).value = f"{self.start_date.minute:02d}"
        else:
            ws1.cell(row=16, column=10).value = ''
            
        ws1.cell(row=16, column=12).value = (
            f" {self.start_date.day:02d}" if (self.start_date) else ""
        )
        ws1.cell(row=16, column=14).value = (
            f" {self.start_date.month:02d}" if (self.start_date) else ""
        )
        
        ws1.cell(row=17, column=8).value = f"{self.end_date.hour:02d}" if (self.end_date) else ''
        if (self.end_date):
            if self.end_date.minute > 0:
                ws1.cell(row=17, column=10).value = f"{self.end_date.minute:02d}"
        else:
            ws1.cell(row=17, column=10).value = ''
            
        ws1.cell(row=17, column=12).value = (
            f" {self.end_date.day:02d}" if (self.end_date) else ""
        )
        ws1.cell(row=17, column=14).value = (
            f" {self.end_date.month:02d}" if (self.end_date) else ""
        )
        # chi huy xe
        # employee_lead_id
        if(self.employee_lead_id):
            ws1.cell(row=19, column=7).value = f"{self.rank_id.name or ''} {self.employee_lead_id.name or ''}"
            ws1.cell(row=19, column=14).value = f"{self.job_id.name or ''}"
        # so km di ve
        ws1.cell(row=21, column=7).value = self.distance_plan
        ws1.cell(row=29, column=1).value = (
            f"{self.employee_command_id.rank_id.name} {self.employee_command_id.name or ''}" if (self.employee_command_id) else ""
        )
        ws1.cell(row=29, column=2).alignment = Alignment(horizontal='center')
        
        ws1.merge_cells(start_row=29, start_column=2, end_row=29, end_column=7)
        
        try:
            if self.employee_approved_id.name:
                ws1.cell(row=29, column=10).value = f"{self.employee_approved_id.rank_id.name} {self.employee_approved_id.name or ''}"
                ws1.cell(row=29, column=10).alignment = Alignment(horizontal='center')

            else:
                ws1.cell(row=29, column=10).value = ""
        except:
            ws1.cell(row=29, column=10).value = ''   
        ws1.merge_cells(start_row=29, start_column=10, end_row=29, end_column=16) 
        
        file_path2 = f'file_command_path2result{self.id}.xlsx'
        workbook.save(file_path2)

        with open(file_path2,"rb") as excel_file:
            file_data = base64.b64encode( excel_file.read())
        # file_data.seek(0)

        # Create an attachment
        attachment = self.env['ir.attachment'].create({
            'name': f'Lenh_Dieu_Phuong_Tien-{self.fleet_command_code:02}.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'fleet.trip',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            # % attachment.id,
            'target': 'new',
        }

    def action_download_template(self):
        # current time
        now = datetime.now()
        # Load the template
        file_path = get_module_resource('fleet_trip', 'static/src/template', 'MY_TEMPLATE.xlsx')
        workbook = openpyxl.load_workbook(file_path)

        # Access the worksheets
        ws1 = workbook['Sheet1']
        # ws1.cell(row=1, column=1).value = self.license_plate
        # start phong dao tao
        ws1.cell(row=3, column=1).value = self.department_plan_id.name.upper()
        # end phong dao tao
        ws1.cell(row=4, column=1).value = f"Số: {self.fleet_code:02}/DTPT-{self.acronym_department_plan}"
        ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4) 
        ws1.cell(row=4, column=5).value = f"Hà Nội, ngày {now.day:02} tháng {now.month:02} Năm {now.year}"
        ws1.cell(row=10, column=7).value = (
            f"Tên phương tiện: {capitalize_first_letter(self.category_plan_name)}"
            if self.category_plan_name
            else "Tên phương tiện: ……...…………."
        )
        ws1.merge_cells(start_row=10, start_column=7, end_row=10, end_column=13) 
        ws1.cell(row=9, column=6).value = (
            f"{self.department_plan_id.name}."
            if self.department_plan_id
            else '-'
        )
        ws1.merge_cells(start_row=9, start_column=6, end_row=9, end_column=10) 
        ws1.cell(row=10, column=1).value = f"Số chuyến (lượt): {self.number_trips:02d}"
        ws1.cell(row=11, column=1).value = (
            f"Nội dung nhiệm vụ: {self.description}"
            if self.description
            else "Nội dung nhiệm vụ: ………………………………………...………"
        )
        ws1.merge_cells(start_row=11, start_column=1, end_row=11, end_column=13) 
        ws1.cell(row=12, column=3).value = f"{self.time_day_compute}"
        ws1.cell(row=14, column=1).value = f"Số người: {self.number_people:02d} người "
        ws1.cell(row=14, column=7).value = f"Số tấn HH: {self.product_weigh} tấn "

        if (self.start_date):
            ws1.cell(row=12, column=5).value = f"Từ: {self.start_date.hour:02d}h "
            ws1.merge_cells(start_row=12, start_column=5, end_row=12, end_column=6) 
            if self.start_date.minute > 0:
                ws1.cell(row=12, column=7).value = f" {self.start_date.minute:02d}"
            ws1.cell(row=12, column=10).value = f" {self.start_date.day:02d} "
            ws1.cell(row=12, column=12).value = f" {self.start_date.month:02d} "
            ws1.cell(row=12, column=13).value = f"năm {self.start_date.year}"
        if (self.end_date):
            ws1.cell(row=13, column=5).value = f"Đến: {self.end_date.hour:02d}h"
            ws1.merge_cells(start_row=13, start_column=5, end_row=13, end_column=6) 
            if self.end_date.minute > 0:
                ws1.cell(row=13, column=7).value = f" {self.end_date.minute:02d}"
            ws1.cell(row=13, column=10).value = f" {self.end_date.day:02d}"
            ws1.cell(row=13, column=12).value = f" {self.end_date.month:02d}"
            ws1.cell(row=13, column=13).value = f"năm {self.end_date.year}"
        # employee_lead_id
        if(self.employee_lead_id):
            ws1.cell(row=15, column=1).value = f"Chỉ huy xe: Họ tên: {capitalize_first_letter(self.employee_lead_id.name) or ''} " +\
                f"C.bậc: {self.rank_id.name or ''} C.vụ: {self.job_id.name or ''}"
            ws1.merge_cells(start_row=15, start_column=1, end_row=15, end_column=13)       
        if (self.location_id):
            ws1.cell(row=16, column=1).value = f"Địa điểm đón xe: {self.location_id.name or '' }," \
                + f" {self.location_id.ward_id.name or ''}," \
                + f" {self.location_id.district_id.name or ''}," \
                + f" {self.location_id.state_id.name or '' },"       
            ws1.merge_cells(start_row=16, start_column=1, end_row=16, end_column=13)     
        if (self.location_dest_id):
            ws1.cell(row=17, column=1).value = f"Nơi đến: {capitalize_first_letter(self.location_dest_id.name) or ''}," \
                + f" {self.location_dest_id.ward_id.name or ''}," \
                + f" {self.location_dest_id.district_id.name or ''}," \
                + f" {self.location_dest_id.state_id.name or ''}"       
            ws1.merge_cells(start_row=17, start_column=1, end_row=17, end_column=13)    
        ws1.cell(row=18, column=1).value = f"Dự kiến tổng số km đi, về (giờ hoạt động) {self.distance_plan} km (giờ)./." 
        ws1.merge_cells(start_row=18, start_column=1, end_row=18, end_column=13)  

        if self.employee_plan_id:
            ws1.cell(row=25, column=1).value = (self.employee_plan_id.name or '')
            ws1.cell(row=25, column=1).alignment = Alignment(horizontal='center')

            if self.employee_plan_id.sign_image:
                image_data = base64.b64decode(self.employee_plan_id.sign_image)
                with open(f'signature{self.employee_plan_id.id}.png', 'wb') as f:
                    f.write(image_data)

                img = Image(f'signature{self.employee_plan_id.id}.png')
                # Insert the image at a specific cell (e.g., B21)

                # ws1.add_image(img, 'B21')
                # Calculate the size of the cells B21 to D25
                cell_width = ws1.column_dimensions['B'].width
                cell_height = ws1.row_dimensions[21].height
                # total_width = cell_width * 3  # B, C, D
                # total_height = cell_height * 4  # 21, 22, 23, 24

                # Convert the size to pixels
                img.width = 125
                img.height = 105

                ws1.add_image(img, 'B21')
                # ws1.add_image(image_data, 'B21')  # Insert the image at cell B2
                # ws1['B21'].value = image_data
                # img = Image(image_data)
                # img.anchor(ws1['B21'])  # Use the cell reference directly
        # command id 
        ws1.merge_cells(start_row=25, start_column=1, end_row=25, end_column=4) 
        try:
            # start nguoi duyet du tru
            approve_user = (f"{self.employee_id.job_id.name} {self.department_plan_id.name}".upper() or '')
            approve_user = remove_duplicate_adjacent_words(approve_user)
            ws1.cell(row=20, column=6).value =  approve_user
            #chu ky
            if self.employee_id.sign_image:
                image_data = base64.b64decode(self.employee_id.sign_image)
                with open(f'signature{self.employee_id.id}.png', 'wb') as f:
                    f.write(image_data)
                img = Image(f'signature{self.employee_id.id}.png')
                # Convert the size to pixels
                img.width = 125
                img.height = 105
                # Insert the image at a specific cell (e.g., B21)
                ws1.add_image(img, 'H21')
            # end nguoi duyet du tru
            ws1.cell(row=25, column=6).value =  (f"{self.employee_id.rank_id.name} {self.employee_id.name}" or '') 
            ws1.cell(row=25, column=6).alignment = Alignment(horizontal='center')
            
        except:
            ws1.cell(row=25, column=6).value = ''   
        ws1.merge_cells(start_row=25, start_column=6, end_row=25, end_column=13)  

        try:
            # có 2 trường hợp văn phòng hoặc hậu cần self.department_id.manager_id.chức vụ, cấp bậc, name
            # chức vụ
            ws1.cell(row=27, column=3).value = self.department_belong_id.manager_id.job_id.name.upper() or ""
            ws1.cell(row=32, column=3).value = (f"{self.department_belong_id.manager_id.rank_id.name} {self.department_belong_id.manager_id.name}" or '')
            ws1.cell(row=32, column=3).alignment = Alignment(horizontal='center')
            
        except:
            ws1.cell(row=32, column=3).value = ''   
        ws1.merge_cells(start_row=32, start_column=3, end_row=32, end_column=11)  

        # ws2 = workbook['Sheet2']

        # # Example data fetching
        # records = self.env['fleet.trip'].search([])

        # # Populate the first worksheet
        # row = 2  # Assuming the first row is for headers
        # for record in records:
        #     ws1.cell(row=row, column=1).value = record.field1
        #     ws1.cell(row=row, column=2).value = record.field2
        #     row += 1

        # # Populate the second worksheet
        # row = 2
        # for record in records:
        #     ws2.cell(row=row, column=1).value = record.field3
        #     ws2.cell(row=row, column=2).value = record.field4
        #     row += 1

        # Save the workbook to a BytesIO object
        # file_data = BytesIO()
        file_path2 = f'file_path2result{self.id}.xlsx'
        workbook.save(file_path2)

        with open(file_path2,"rb") as excel_file:
            file_data = base64.b64encode( excel_file.read())
        # file_data.seek(0)

        # Create an attachment
        attachment = self.env['ir.attachment'].create({
            'name': f'Du_tru_phuong_tien-{self.acronym_department_plan}-{self.fleet_code:02}.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'fleet.trip',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            # % attachment.id,
            'target': 'new',
        }


class StockDelvery(models.Model):
    _name = 'stock.delivery'
    _rec_name = 'code'
    _description = 'Phiếu xuất kho'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    code = fields.Char(string='Số phiếu', required=True)
    project_id = fields.Many2one('fleet.project', string='Dự án', required=True)
    category_id = fields.Many2one('fleet.category', string='Hạng mục')
    stock_date = fields.Date(string="Ngày", default=fields.Date.today)
    location_dest_id = fields.Many2one('fleet.location', 'Điểm đích', required=True)
    partner_receive_id = fields.Many2one('res.partner', string='Người nhận', required=True)
    partner_receive_phone = fields.Char(related='partner_receive_id.phone', string='Điện thoại')
    shipping_id = fields.Many2one('res.partner', string='Đơn vị vận chuyển', required=True)
    driver_id = fields.Many2one('res.partner', string='Lái xe', required=True)
    driver_phone = fields.Char(related='driver_id.phone', string='Điện thoại')
    equipment_id = fields.Many2one('maintenance.equipment', string='Xe', required=True)
    delivery_line = fields.One2many('stock.delivery.line', 'delivery_id', string='Chi tiết xuất kho')


class StockDelveryLine(models.Model):
    _name = 'stock.delivery.line'
    _description = 'Chi tiết xuất kho'

    delivery_id = fields.Many2one('stock.delivery', string='Phiếu xuất kho')
    product_id = fields.Many2one('product.template', string='Sản phẩm', required=True)
    section = fields.Char(related='product_id.section', string='Tiết diện')
    product_length = fields.Integer(related='product_id.product_length', string='Dài')
    uom_id = fields.Many2one(related='product_id.uom_id', string='Đơn vị')
    out_qty = fields.Float(string='SL Xuất')
    bao_qty = fields.Float(string='Bao')
    note = fields.Text(string='Ghi chú')
